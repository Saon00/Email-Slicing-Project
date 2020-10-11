from tkinter import *
from openpyxl import *

wb = load_workbook('E:\Python\Youtube\EmailSlice.xlsx')  # file or file path, you should change this according to you
sheet = wb.active


def excel_file():
    # resizing the width of columns in sheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30

    # Heading the Value Name
    sheet.cell(row=1, column=1).value = "E-mail name"
    sheet.cell(row=1, column=2).value = "User name"
    sheet.cell(row=1, column=3).value = "Domain name"


def clear_text():
    input_field.delete(0, END)  # for multiple uses


def clickme():
    temp_text = input_field.get()
    user_name = temp_text[:temp_text.index("@")]
    domain_name = temp_text[temp_text.index("@") + 1:]
    result = f"""{temp_text}             {user_name}                {domain_name}     

"""  # formatting result or output

    output.insert(END, result)

    if (input_field.get() == ""):
        print("Empty Input")
    else:
        current_row = sheet.max_row
        current_column = sheet.max_column   # optional use

        sheet.cell(row=current_row + 1, column=1).value = input_field.get()
        sheet.cell(row=current_row + 1, column=2).value = user_name
        sheet.cell(row=current_row + 1, column=3).value = domain_name

        wb.save("E:\Python\Youtube\emailSlicing2.xlsx")  # file or file path, you should change this according to you
        clear_text()

# Driver Code
if __name__ == '__main__':
    window = Tk()
    window.title("E-Mail Slicing Project")
    window.geometry('800x500')

    # calling excel sheet
    excel_file()

    # Top level part
    frame1 = Label(window, text="E-mail Slicing", font=("Arial", 20), bg="#080808", fg="white", width=500, height=2)
    frame1.pack()

    # Left Sentence
    label1 = Label(window, text="Write an E-mail: ", font=("Arial", 15), fg="#010847")
    label1.pack(pady=25)  # label1.place(x=10, y=100)

    # as our input will be String type
    email_name = StringVar()

    # User input
    input_field = Entry(window, font=("Arial", 20), textvariable=email_name, bg='#8dc94d', width=25)
    input_field.pack()  # input_field.place(x=180, y=100)

    # create Button
    btn1 = Button(window, text="Click Here to Slice e-mail", bg="#8591de", font=('Bold', 10), command=clickme)
    btn1.pack(pady=40)

    # Output box
    output = Text(window, width=100, height=15, bg="#c195e6", fg='black', font=('Ariel', 15))
    output.pack()


    # exit button
    btn2 = Button(window, text="Exit the Program", bg="#94140d", fg="white", font=('Bold', 10), command=window.destroy)
    btn2.pack(pady=20)

    window.mainloop()
